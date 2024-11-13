import streamlit as st
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from datetime import datetime
import os
import io

# Excel dosyasına veriyi kaydetme ve sertifika numarasına köprü ekleme
def save_to_excel(data, image_folder):
    excel_file = io.BytesIO()
    workbook = Workbook()

    # Sertifikalar sayfası
    cert_sheet = workbook.active
    cert_sheet.title = "Sertifikalar"
    
    # Başlıkları ekle
    headers = ['Ürün Tanımı', 'Kalite', 'Firma', 'Sertifika No', 'Eklenme Tarihi', 'Sertifika Fotoğrafı']
    cert_sheet.append(headers)

    # Veriyi ekle
    for i, item in enumerate(data, start=2):  # 2'den başlıyoruz, çünkü 1. satır başlık
        added_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        item.append(added_date)  # Eklenme tarihini ekliyoruz

        # Sertifika fotoğrafı yolunu ekle
        cert_image_path = os.path.join(image_folder, f"{item[3]}.jpg")  # Sertifika No'ya göre fotoğraf ismi
        item.append(cert_image_path)

        # Satırda veriyi ekle
        for j, value in enumerate(item[:6], start=1):  # 6 sütun olacak, fotoğraf için ekledik
            cert_sheet.cell(row=i, column=j, value=value)

        # Sertifika No'ya köprü ekle
        cert_sheet[f'D{i}'].hyperlink = cert_image_path
        cert_sheet[f'D{i}'].style = 'Hyperlink'  # Hyperlink stilini uygula

    # Sertifika Fotoğrafı sayfası (resimler)
    img_sheet = workbook.create_sheet(title="Sertifika Fotoğrafı")
    for i, item in enumerate(data, start=1):
        cert_image_path = os.path.join(image_folder, f"{item[3]}.jpg")  # Sertifika numarası ile resim yolu
        if os.path.exists(cert_image_path):
            img = Image(cert_image_path)
            img_sheet.add_image(img, f'A{i}')  # Resmi ekle (A sütununa)

    # Excel dosyasını byte dizisi olarak kaydet
    workbook.save(excel_file)
    excel_file.seek(0)  # Dosya başına geri dön
    return excel_file

# Streamlit arayüzü
st.title('Duyar Metal Kalite Sertifikaları Yönetim Sistemi')

# Veri ekleme kısmı
st.header('Yeni Ürün Ekle')

urun_tanim = st.text_input('Ürün Tanımı:')
kalite = st.text_input('Kalite:')
firma = st.text_input('Firma Adı:')
sertifika_no = st.text_input('Sertifika No:')
sertifika_resmi = st.file_uploader("Sertifika Fotoğrafı Yükle", type=["jpg", "jpeg", "png"])

# Sertifika fotoğrafı yüklenmişse, dosya sistemine kaydedelim
if sertifika_resmi:
    image_folder = '/tmp/SertifikaFotoğrafları'  # Geçici dizine kaydedilecek
    os.makedirs(image_folder, exist_ok=True)
    image_path = os.path.join(image_folder, f"{sertifika_no}.jpg")
    with open(image_path, "wb") as f:
        f.write(sertifika_resmi.getbuffer())

# Yeni ürün eklemek için buton
if st.button('Ürün Ekle'):
    if urun_tanim and kalite and firma and sertifika_no:
        # Yeni veriyi Excel dosyasına ekleyelim
        new_data = [[urun_tanim, kalite, firma, sertifika_no]]  # Sadece gerekli veriler eklendi
        excel_file = save_to_excel(new_data, image_folder)

        # Kullanıcıya Excel dosyasını indirme seçeneği sunalım
        st.download_button(
            label="Excel Dosyasını İndir",
            data=excel_file,
            file_name='Sertifika_Kayitlari.xlsx',
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        st.success('Yeni ürün başarıyla eklendi ve Excel dosyasına kaydedildi.')
    else:
        st.error('Lütfen tüm alanları doldurun!')

# Arama / Filtreleme Kısmı
st.header('Verileri Ara')

# Excel dosyasını oku
excel_file = '/tmp/Sertifika_Kayitlari.xlsx'  # Geçici dosya yolu
if os.path.exists(excel_file):
    df = pd.read_excel(excel_file)

    # Filtreleme seçenekleri ekleyelim
    search_term = st.text_input('Arama Yapın (Ürün, Kalite, Firma, Sertifika No)')
    if search_term:
        filtered_df = df[df.apply(lambda row: row.astype(str).str.contains(search_term, case=False).any(), axis=1)]
    else:
        filtered_df = df

    # Filtrelenmiş veriyi göster
    st.dataframe(filtered_df)

else:
    st.warning("Henüz veri eklenmedi.")
